#Pic -> SpreadSheet
import string
import xlsxwriter
from openpyxl import load_workbook      # <- imports
from openpyxl.styles import PatternFill
from PIL import Image
import csv


times = 0

while times < 10:
  times += 1
  try:
    Image_name = input("Please input the path of your image file here: ")
    times = 10
  except Exception as e:
    print(F"REAL ERROR {e}\n\nDev note: the file path may be incorrect, the file extension may be wrong, the file may not exist")

def Write_File(A1,A2,A3):
  with open("Test.txt","w")as file:

    file.write("\n\nFIRST ARRAY \n"+ str(A1))
    file.write("\n\nSECOND ARRAY \n"+ str(A2))
    file.write("\n\nTHIRD ARRAY \n"+ str(A3))

def Den2HexCol(R=False,G=False,B=False, value=0):      #<- takes RGB values in denary and converts it to its hex counterpart
  if value != None:

    if value <= 15 and value >=3 :
      value = 20
  
    string = ""
    #print(value)
    value = hex(value)
    value = value.split('x')

    if R:
      string = f"{value[1]}0000"
    elif G:
      string = f"00{value[1]}00"
    elif B:
      string = f"0000{value[1]}"


    if len(string)<= 5:
      if R:
        string += '0'
      elif G:
        string += '0'
      elif B:
        string = '0' + string
      
    return string
  


print("Processing...")

FILENAME = 'Image_data.xlsx' #Excel file to be written and read


image = Image.open(f'{Image_name}') # <- the Image in question being loaded with PIL
pixels = image.load() #<- Loading the pixels into a data structure (2D array), 1st element = pixel, 2nd element = pixel information

workbook = xlsxwriter.Workbook(f'{FILENAME}') # Opening the excel workbook in xlsxwriter
worksheet = workbook.add_worksheet()  #adding a worksheet to the workbook

excel = list(string.ascii_uppercase)  #Array of the alphabet in uppercase
x = len(excel)  #Current length of unmodified excel array, this is so the for loop doesnt go on for infinity

for a in range(x):
  for b in range(x):
    excel.append(f"{excel[a]}{excel[b]}")     #Appends more Excel formatted letters AA, AB , AC ect

for a in range(x):
  for b in range(x):  
    for c in range(x):  
      excel.append(f"{excel[a]}{excel[b]}{excel[c]}")
a = 0
b = 0   # <- these numbers are used to determine what row a colour is on, a = r, b= g, c = g
c = 0
End_Array1 = [] #<- these arrays are all the Excel locations End_Array1 = all the r values on the spreadsheet
End_Array2 = []
End_Array3 = []
f = 0

for col in range(image.size[0]):
  if f >= 3:
    f = 0         # F can be used RGB, 1 = r, 2 = g , 3 = b

  f += 1
  a += 1
  b += 1
  c += 1
  #print(f)  
  for row in range(image.size[1]):      #TODO write a way to split RGB up into 3 sections //DONE

      if f == 1:
        #print(f'{excel[row-1]}{a}')
        worksheet.write(col,row,pixels[col,row][0])
        End_Array1.append(f'{excel[row]}{a}')
        #print(f'{excel[row]}{a}')
        
      elif f == 2:
        worksheet.write(col, row,pixels[col,row][1])
        End_Array2.append(f'{excel[row]}{a}')
      elif f== 3:
        worksheet.write(col, row,pixels[col,row][2])
        End_Array3.append(f'{excel[row]}{a}')


workbook.close() # END OF XLSXWriter to ensure theres no module confliction

wb = load_workbook(filename = f'{FILENAME}')
ws = wb.active
sheet = wb['Sheet1']



for i in range(len(End_Array1)-1):
  #print(sheet[End_Array2[i]].value)
  conv = Den2HexCol(True,False,False,sheet[End_Array1[i]].value)
  #print(conv)
  try:
    sheet[End_Array1[i]].fill = PatternFill(start_color=conv,fill_type = "solid")
  except Exception as e:
    print(f"ERROR CONV IS: {conv}")



for i in range(len(End_Array2)-1):
  #print(sheet[End_Array2[i]].value)
  conv = Den2HexCol(False,True,False,sheet[End_Array2[i]].value)
  #print(conv)
  try:
    sheet[End_Array2[i]].fill = PatternFill(start_color=conv,fill_type = "solid")
  except Exception as e:
    print(f"ERROR CONV IS: {conv}")


for i in range(len(End_Array3)-1):
  #print(sheet[End_Array2[i]].value)
  conv = Den2HexCol(False,False,True,sheet[End_Array2[i]].value)
  #print(conv)
  try:
    sheet[End_Array3[i]].fill = PatternFill(start_color=conv,fill_type = "solid")
  except Exception as e:
    print(f"ERROR CONV IS: {conv}")








wb.save(f'{FILENAME}')



